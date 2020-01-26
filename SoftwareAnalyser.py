#Packages needed
import re
import regex
import pandas as pd
import numpy as np
import networkx as nx
import inspect
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import time
import sys
import numpy as np
import sqlite3
import openpyxl as ex
from random import randint
from datetime import datetime

class Database:
    #Database class which deals with saving all files and the sql table
    def __init__(self,Excel):

        #### Objective 18,19

        #### Relational Database

        #Creates overall table of usernames and their preferences if it does not yet exist
	   # Table only created if it does not yet exist
        self.conn = sqlite3.connect('Storage.db')
        stmt = """CREATE TABLE IF NOT EXISTS PersonProperties
                (ID CHAR(50) PRIMARY KEY     NOT NULL,
                 COMMENTS           INT    NOT NULL,
                 LINES            INT     NOT NULL,
                 ARGUMENTS        INT NOT NULL);"""
        self.conn.execute(stmt)
        book = ex.load_workbook(Excel)

        #Excel and graph tables initialized with their ID as the foregin KEY
        #Of the PersonProperties table
        stmt = """CREATE TABLE IF NOT EXISTS EXCEL
                    (ID CHAR(50) NOT NULL,
                    FILELOCATION CHAR(50) NOT NULL,
                    FOREIGN KEY(ID) REFERENCES PersonProperties(ID));"""

        self.conn.execute(stmt)

        stmt = """CREATE TABLE IF NOT EXISTS GRAPH
                    (ID CHAR(50) NOT NULL,
                    FILELOCATION CHAR(50) NOT NULL,
                    FOREIGN KEY(ID) REFERENCES PersonProperties(ID));"""

        self.conn.execute(stmt)
        #Inserts the persons preferences if they are in the Excel spreadsheet

        self.name = None
        if "Data" in book:
            sheet = book["Data"]
            Name = sheet.cell(row = 2,column = 1).value
            self.name = Name

            NumLines = sheet.cell(row = 2,column = 2).value
            NumComments = sheet.cell(row = 2,column = 3).value
            NumParameters = sheet.cell(row = 2,column = 4).value
            DatabaseName = randint(100,1000)
            query = """SELECT ID
                    FROM PersonProperties;"""
            result = self.conn.execute(query)
            result = [i[0] for i in result]
            if self.name not in result:
                stmt = f"""INSERT INTO PersonProperties(ID,COMMENTS,LINES,ARGUMENTS)
                            VALUES('{self.name}',{NumComments},{NumLines},{NumParameters})"""
                self.conn.execute(stmt)
                self.conn.commit()

        self.getPreferences()

    def hasPreferences(self):
        if self.name is None:
            return False
        query = """SELECT ID
                    FROM PersonProperties;"""
        result = self.conn.execute(query)
        result = [i[0] for i in result]
        if self.name in result:
            return True
        return False

    def getPreferences(self):
        query = """SELECT * FROM PersonProperties;"""

        result = self.conn.execute(query)
        for item in result:
            if item[0] == self.name:
                return item


    def saveExcel(self,Excel,saveLocation):

        #### Objective 11,23

        #Unique timestamp produced
        now = datetime.now()
        book = ex.load_workbook(Excel)
        #unique timestamp placed at the end of the name
        book.save(saveLocation + "/" + (str(now).replace(":",";")).split(".")[0] +".xlsx")
        loc = saveLocation + "/" + (str(now).replace(":",";")).split(".")[0] +".xlsx"
        #location saved in the database


        if self.name is not None:
            stmt = f''' INSERT INTO EXCEL(ID, FILELOCATION)
                     VALUES('{self.name}','{loc}') '''

            self.conn.execute(stmt)
            self.conn.commit()


    def saveGraph(self,DG,Name,location):
        #same process for the graph as the spreadsheet just different file format
        now = datetime.now()
        timeAdd = (str(now).replace(":",";")).split(".")[0]
        completeName = location + "/" + Name + timeAdd + "." + "graphml"
        nx.write_graphml(DG,completeName)


        if self.name is not None:
             stmt = f''' INSERT INTO GRAPH (ID,FILELOCATION)
                   VALUES('{self.name}','{completeName}') '''

             self.conn.execute(stmt)
             self.conn.commit()




    def Close(self):
        self.conn.close()




class Main():

    def __init__(self,CodeFile,saveLocation,Excel): # need to implement Excel = None
        if Excel == None:
            wb = ex.Workbook()
            wb.create_sheet("Sheet1")
            wb.save(filename = 'AnalysisResults.xlsx')
            Excel = "AnalysisResults.xlsx"
        #this called when run analysis button pressed on GUI
        #first the database class is created
        self.database = Database(Excel)
        #Function to take in the script and convert it to text
        self.TakeInScript(CodeFile)
        #static analyses class is called and given inputs to use
        Static = StaticAnalyses(self.script,saveLocation,Excel,self.database)
        #function classes are obtained so that they can be passed to the dynamic analysis class
        funcs = Static.getArrFuncs()
        #Dynamic analysis class started with inputs given to run dynamic analysis
        Dynamic = DynamicAnalyses(self.script,Excel,funcs,saveLocation,self.database)
        #Final version after all editing done by different classes complete and database
        #is then used to save it to the right location
        self.database.saveExcel(Excel,saveLocation)
        #all edits to database are saved and closed
        self.database.Close()

    def TakeInScript(self,location):
        #Function that takes in the script and converts it to text
        self.script = open(location,"r").read()



#Global function to remove comments as diffrerent classes need it
def removeComments(scriptorig):
    script = ""
    #all lines are looped through and only lines that
    #dont start with a # are then added back in to the
    #string which is then returned, if line contains
    #code and comment then hastag is found and everything
    #after is stripped from the script
    for line in scriptorig.splitlines():
        #obtain where first charachter is as
        #there could be blankspaces first
        commentstarts = len(line)-1
        if commentstarts == -1:
            continue
        #locate where/if hashtag is
        for index,let in enumerate(list(line)):
            if let == "#":
                commentstarts = index-1
                break
        #script appended to final script to be returned
        #with anything after the hashtag removed
        script = script + line[:commentstarts+1] + "\n"
    return script

class StaticAnalyses():
    #Orchestrates all the static analyses
    # data collected = self.NumComments , self.NumLines,
    def __init__(self,script,saveLocation,Excel,Database):
        self.database = Database
        self.saveLocation = saveLocation
        self.script = script
        self.scriptNoComments = removeComments(self.script)
        #Create all function classes
        self.CreateFuncs()
        #Create the statc graph
        self.CreateNetGraph()
        #Get number of lines
        self.NumLines()
        #Get number of comments
        self.NumComments()
        #Save info to spreadsheet which has be found out
        self.WriteUpInfo(Excel)

    def WriteUpInfo(self,Excel):

        #### Objective 10,20

        #Excel spreadsheet loaded
        book = ex.load_workbook(Excel)
        #New sheet to save info created


        try:
            sheet2 = book["Static Data"]
        except:
            book.create_sheet("Static Data")
            sheet2 = book["Static Data"]
        #Headings created
        NumLines = 10000000000
        NumComments = 100000000000
        NumParameters = 10000000000000000
        redFill = ex.styles.PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        if self.database.hasPreferences():
            name,NumLines,NumComments,NumParameters = self.database.getPreferences()
        sheet2.cell(row = 1,column = 1).value = "Num Lines"
        NLCELL = sheet2.cell(row = 2,column = 1)
        NLCELL.value = self.NumLines
        if NLCELL.value > NumLines:
            NLCELL.fill = redFill
        sheet2.cell(row = 1,column = 2).value = "Num Comments"
        NCCELL = sheet2.cell(row = 2,column = 2)
        NCCELL.value = self.NumComments
        if NCCELL.value > NumComments:
            NCCELL.fill = redFill
        sheet2.cell(row = 1,column = 3).value = "Num Functions"
        sheet2.cell(row = 2,column = 3).value = len(self.Funcs)
        sheet2.cell(row = 1,column = 5).value = "Functions"
        sheet2.cell(row = 1,column = 6).value = "Num Arguments"
        sheet2.cell(row = 1,column = 7).value = "NumLocalVars"
        sheet2.cell(row = 1,column = 8).value = "StackSize"
        sheet2.cell(row = 1,column = 9).value = "Var Names"
        sheet2.cell(row = 1,column = 10).value = "Num Lines"
        counterRow = 2
        counterCol = 5
        #For each property of each function information saved
        for func in self.Funcs:
            sheet2.cell(row = counterRow,column = counterCol).value = func.getName()
            counterRow += 1
        counterRow = 2
        counterCol += 1
        for func in self.Funcs:
            sheet2.cell(row = counterRow,column = counterCol).value = func.FuncInfoDict["NumArguments"]
            if sheet2.cell(row = counterRow,column = counterCol).value > NumParameters:
                sheet2.cell(row = counterRow,column = counterCol).fill = redFill
            counterRow += 1
        counterRow = 2
        counterCol += 1
        for func in self.Funcs:
            sheet2.cell(row = counterRow,column = counterCol).value = func.FuncInfoDict["NumLocalVars"]
            counterRow += 1
        counterRow = 2
        counterCol += 1
        for func in self.Funcs:
            sheet2.cell(row = counterRow,column = counterCol).value = func.FuncInfoDict["StackSize"]
            counterRow += 1
        counterRow = 2
        counterCol += 1
        for func in self.Funcs:
            sheet2.cell(row = counterRow,column = counterCol).value = str(func.FuncInfoDict["varNames"])
            counterRow += 1
        counterRow = 2
        counterCol += 1
        for func in self.Funcs:
            sheet2.cell(row = counterRow,column = counterCol).value = len(func.FuncText.splitlines())
            counterRow += 1
        #spreadsheet saved
        book.save(Excel)



    def getArrFuncs(self):
        #Function to encapsulate the self.Funcs array
        return self.Funcs



    def CreateNetGraph(self):

        #### Objective 21,22

        #Function to convert the graph made up of Functions classes
        #To a networkX graph which can be saved in the graphml format
        DG = nx.DiGraph()
        #Adds all functions as nodes to the graph
        for Func in self.Funcs:
            DG.add_node(Func.getName())
        #add all edges by double looping through functions and their children
        for Func in self.Funcs:
            for child in Func.getArrFuncsCalled():
                DG.add_edges_from([(Func.getName(),child.getName())])
        #Graph saved through database class
        self.database.saveGraph(DG,"StaticFuncInteracting",self.saveLocation)



    def NumLines(self):

        #### Objective 4

        #Calculate number of lines by splitting script
        #into a list with a line per item and using len function
        self.NumLines = len((self.script.splitlines()))


    def NumComments(self):

        #### Objective 5

        #looping through each line and adding one
        #if there is a hashtag in the line
        self.NumComments = 0
        #Loop through all the lines in the script
        for line in self.script.splitlines():
            #Check if line has a hashtag
            if "#" in line:
                #Initialize variables such that no quotes have
                #yet been opened
                OpenSingle = 0
                OpenDouble = 0
                Correct = 0
                #Loop through all charachters in the line
                for char in line:
                    #If it is a single quote and it is not a
                    #open double quote then toggle the single quote variable
                    if char == "'" and OpenDouble == 0:
                        OpenSingle += 1
                        OpenSingle %= 2
                    #If it is a double quote and there is not a
                    #open single quote then toggle the double quote variable
                    if char == '"' and OpenSingle == 0:
                        OpenDouble += 1
                        OpenDouble %= 2
                    #If it is a # and not within any quotes then there is
                    #a comment
                    if char == "#" and OpenDouble == 0 and OpenSingle == 0:
                        Correct = 1
                    #Loop through rest of the line as there could be a valid
                    #and invalid hashtag symbol
                #If there was a single valid hastag then add one to the number
                #of comments
                if Correct == 1:
                    self.NumComments += 1





    def CreateFuncs(self):

        #### Objective 6

        #### Aggregation

        #Creates a Function Object for every function in the script
        self.Funcs = []
        for index,line in enumerate(self.script.splitlines()):
            line2 = line.strip()
            #finds if first three charachters in a line is "def"
            #Also works for functions in functions as it ignores indentations
            # "def " used at the start of line used to make sure it definitely a function definition at the
            #the line[0:4] checks it is at the start of the line as well. These two checks make sure only function definitions
            #are found
            if line2[0:4] == "def ":
                #creates a function class adding it to is self.Funcs array
                # giving the function the whole script and the line thats its defenition is on
                self.Funcs.append(Function(self.script,index,self.scriptNoComments))
        #Calls each function object to find out data about themselves
        #Before each function can point to other function they all have to know their names
        #Therefore FindName() funcion called first
        for Func in self.Funcs:
            Func.FindName()
        #Then for all function data collecting function run to get information
        for Func in self.Funcs:
            Func.collectData(self.Funcs)




class Function():
    #Class for every function in the script
    def __init__(self,script,line,scriptNoComments):
        self.scriptNoComments = scriptNoComments
        self.script = script
        self.line = line
        self.scriptLines = self.script.splitlines()

    def collectData(self,otherFuncs):
        self.Defenition = self.scriptLines[self.line]
        self.otherFuncs = otherFuncs
        #works out what from the script is the functions text
        self.FuncText = self.getFuncText()
        #finds out what other functions are called
        self.GetFuncsCalled()
        #Get dynamic metrics by executing the Function
        self.FuncInfoDict = self.executeFunctionAndGatherInfo(self.FuncText)
        self.FuncInfoDict["Name"] = self.getName()

    def getArrFuncsCalled(self):
        return self.FuncsCalled

    def executeFunctionAndGatherInfo(self,FuncText):

        #### Objective 7,8,9

        #This function gets information by creating the function within the
        #class and then dynamically getting metrics on it such as size taken up

        #function is executed
        exec(FuncText,globals())
        codeInfos = {"NumArguments":"argcount","NumLocalVars":"nlocals","StackSize":"stacksize","varNames":"varnames"}
        FuncInfoDict = {}
        for info in codeInfos:
            #Certain property of the function found out
            strEx =  info + " = "+self.getName()+".__code__.co_" + codeInfos[info]
            exec(strEx,globals(),FuncInfoDict)

        return FuncInfoDict

    def getLineNums(self):
        #gives range of lines that function in so that code can decide what line is in which function
        return (self.line,self.line + len(self.FuncText.splitlines())-1)


    def getNumParameter(self):
        #Encapsulator for num arguements item within FuncInfoDict
        return self.FuncInfoDict["NumArguments"]



    def GetFuncsCalled(self):
        #Function to find out which functions are called  by this function
        self.FuncsCalled = set()
        #Remove comments so easier to analyse
        self.FuncTextNoCommas = removeComments(self.FuncText)
        self.FuncLines = self.FuncText.splitlines()
        FuncLines = self.FuncLines[1::]
        names = [Func.getName() for Func in self.otherFuncs]

        #Creates a regex statement to OR look for any function name
        orString = "("
        for item in names:
            orString = orString + item + "|"
        orString = orString[:-1]
        orString = orString + ")"
        OtherFuncCalls = []
        for line in FuncLines:
            #Regex run that is looking for any of the function names within this functions text
            matches = re.finditer(orString+r'\(.*\)',line)
            #Looping through all matches found
            for i in matches:
                OtherFuncCalls.append(i.group(0))
                name = i.group(0).split("(")[0]
                #function added to the FuncsCalled array by using the getFuncFromName function
                #so that can point to actual class and not just the name
                self.FuncsCalled.add(self.getFuncFromName(name,self.otherFuncs))


    def getFuncFromName(self,Name,Funcs):
        #Allows a function Class to be found from its name

        #Loop through all function to find out which one has the same name
        for Func in Funcs:
            if Func.getName() == Name:
                return Func
        #Something has gone wrong if no func has been returned
        #and therefore error raised
        raise AssertionError("Func name can't be found")





    def getFuncText(self):
        #function to work out where the text of the function finishes

        #First bit works out how many spaces before def in the first line
        #of the function and therefore you know how many spaces you need before
        #a charachter at the end of a function to know that the function text
        #is finished
        firstLine = list(self.scriptLines[self.line+1])
        for index,let in enumerate(firstLine):
            min = index
            if let != " ":
                break
        finalLine = -1
        #now looping through text looking for where spaces before charachter
        #is smaller than or equal to one set up in earlier parr of the code
        for index,line in enumerate(self.scriptLines[self.line+1::]):
            for index2,let in enumerate(list(line)):
                numspaces = index2
                if let != " ":
                    break
            #this checking for smaller then breaking as finalLine has been found
            if numspaces < min:
                finalLine = index + 1
                break
        #If no finalLine found the function goes until the end of the code
        if finalLine == -1:
            self.FuncText = "\n".join(self.scriptLines[self.line::])
        else:
            #Otherwise function text is from the start until finalLine
            self.FuncText = "\n".join(self.scriptLines[self.line:finalLine+self.line])
        firstLine = list(self.scriptLines[self.line])
        for index,let in enumerate(firstLine):
            min = index
            if let != " ":
                break
        Function = ""
        #Joining back funciton text together and returning
        for line in self.FuncText.splitlines():
            Function = Function + line[min::] + "\n"
        return Function ###Check This


    def FindName(self):
        #Finds the name by splitting the definition
        #and working out the text after def and before the (
        self.Defenition = self.scriptLines[self.line]
        self.name =  self.Defenition.split()[1].split("(")[0]

    def getName(self):
        return self.name


class Tree:
    #This is the graph class
    def __init__(self):
        self.nodes = []

    def isCyclicSpecific(self,node,vis,stack):

        #### Depth First Search & Recursion

        #this node has now been visited and is placed on the recursion stack
        vis[node] = True
        stack[node] = True
        #look at all forward connected neighbours
        for neighbour in node.connectedToForwards():
            #if neighbour been visited do not need to check
            if not vis[neighbour]:
                if self.isCyclicSpecific(neighbour,vis,stack):
                    return True
            elif stack[neighbour]:
                return True
        stack[node] = False
        return False
    def isCyclic(self):

        #### Objective 2

        #have a visited dictionary to keep track of  which nodes have been visited
        vis = {node:False for node in self.nodes}
        stack = {node:False for node in self.nodes}
        #loop through all the nodes
        for node in self.nodes:
            #Only need to check if havent been visited
            if not vis[node]:
                if self.isCyclicSpecific(node,vis,stack):
                    return True
        #If no cyclicness found return False
        return False



    def cyclomComplex(self,start):

        #### Objective 29

        #### Breadth First Search & Recursion

        numPaths = 0
        done1 = 0
        for edge in start.edges:
            if edge.node2 != start:
                done1 = 1
                numPaths += self.cyclomComplex(edge.node2)
        if done1 == 0:
            numPaths += 1
        return numPaths



    def save(self,saveLocation,Database):
        #This function used to convert graph to networkX graph so that it can be saved
        #as a graphml file
        DG = nx.DiGraph()
        #double loop through node and children
        for node in self.nodes:
            for child in node.connectedTo():
                edge = node.getEdge(child)
                if edge.node1 == node:
                    #weighted edge added and weight of node put as text under name so that it can be viewed in the GUI
                    DG.add_weighted_edges_from([(node.getName() + '\n' + str(node.weight),child.getName() + '\n' + str(child.weight),edge.weight)])
        Database.saveGraph(DG,"ReducedDynamicFuncInteracting",saveLocation)
        #nx.write_graphml(DG,saveLocation + "/ReducedDynamicFuncInteracting.graphml")

    def createNode(self,node,weight):
        #node createc with a certain weight
        self.nodes.append(Node(node,weight))

    def createPath(self,node1,weight1,node2,weight2,edgeWeight):
        #weighted edge creation nodes also created if
        #they dont exist
        if not self.inTree(node1):
            self.nodes.append(Node(node1,weight1))
        if not self.inTree(node2):
            self.nodes.append(Node(node2,weight2))
        self.createEdge(node1,node2,edgeWeight)

    def createEdge(self,node1,node2,edge):
        #Edge class created between two nodes with its weight
        node1 = self.getNode(node1)
        node2 = self.getNode(node2)
        edge = Edge(node1,node2,edge)
        node1.addEdge(edge)
        node2.addEdge(edge)

    def inTree(self,nodeN):
        #Check if a node is in a tree and either returning
        #True or False
        for node in self.nodes:
            if node.getName() == nodeN:
                return True
        return False

    def getNode(self,nodeN):
        #Node object based on name is found and returned
        if nodeN == "module":
            for node in self.nodes:
                if "module" in node.getName():
                    return node
        for node in self.nodes:
            if node.getName() == nodeN:
                return node

    def reduceOnce(self):
        #One pair that can be reduced is found reduced and then function finishes
        #If no pair found True is returned to show that graph is fully reduced

        #Double loop through nodes and their children
        for node in self.nodes:
            for child in node.connectedTo():
                #Check to see both node,child have 2 or less nodes
                #which is a requirement for compressing
                if node.numChildren() <= 2 and child.numChildren() <= 2:
                    #If they both only have one node they can always be combines
                    #irrespective of the direction
                    if node.numChildren() == 1 and child.numChildren() == 1:
                        #Combine function of the two nodes and then exit with False
                        #as other pair could still exist
                        self.combine(node,child)
                        return False
                    #if node has one, of course if child has then this will be checked later on in the loop
                    #at this point if node has 1 child then the child must have 2
                    if node.numChildren() == 1:
                        #OtherEdge and connectorEdge located of the child
                        for edge in child.edges:
                            if node in edge.nodes and child in edge.nodes:
                                connectorEdge = edge
                            else:
                                otherEdge = edge

                        #Check to see direction is in one way as described in the NEA writeup
                        if otherEdge.node2 == child and connectorEdge.node2 == node or otherEdge.node1 == child and connectorEdge.node1 == node:
                              #Combined
                              self.combine(node,child)
                              return False
                    #Final scenario if they both have 2
                    if node.numChildren() == 2 and child.numChildren() == 2:
                        #otherEdge and connectorEdge must be located for both node and child
                        for edge in child.edges:
                            if node in edge.nodes and child in edge.nodes:
                                connectorEdge = edge
                            else:
                                otherEdge = edge
                        for edge in node.edges:
                            if node in edge.nodes and child in edge.nodes:
                                connectorEdge2 = edge
                            else:
                                otherEdge2 = edge
                        #check if direction all in the same way
                        if otherEdge.node2 == child and connectorEdge.node2 == node and otherEdge2.node1 == node or otherEdge.node1 == child and connectorEdge.node1 == node and otherEdge2.node2 == node:
                            self.combine(node,child)
                            return False

        #If got to here no pairs found so graph fully reduced
        return True

    def reduce(self):

        #### Objective 28

        #Function to reduce works by repeatedly calling reduceOnce()
        #until it returns True showing graph fully reduced
        while not self.reduceOnce():
            #any edges that now dont have nodes on both sides becuase
            #one side has been combined are deleted
            for node in self.nodes:
                node.removeEdgesWasted(self.nodes)

    def combine(self,node1,node2):
        #Function to combine two nodes as they have been found combinable


        #first two nodes are removed from the nodes array
        self.nodes.remove(node1)
        self.nodes.remove(node2)
        #name of new node created by putting a ',' in between the two names
        newName = node1.getName() + "," + node2.getName()
        #combined node created
        self.createNode(newName,node1.weight+node2.weight)
        combinedNode = self.getNode(newName)
        #edge in original node1 that combines node1 with node2 is removed
        #this also happens with node2
        for edge in node1.edges:
            if node2 in edge.nodes:
                node1.edges.remove(edge)
        for edge in node2.edges:
            if node1 in edge.nodes:
                node2.edges.remove(edge)
        #edges in node1 and node2 updated so that instead of pointing to node1
        #or node2 they point to the combined node
        for edge in node1.edges:
            if edge.node1 == node1:
                edge.node1 = combinedNode
                edge.nodes[0] = combinedNode
            if edge.node2 == node1:
                edge.node2 = combinedNode
                edge.nodes[1] = combinedNode
        for edge in node2.edges:
            if edge.node1 == node2:
                edge.node1 = combinedNode
                edge.nodes[0] = combinedNode
            if edge.node2 == node2:
                edge.node2 = combinedNode
                edge.nodes[1] = combinedNode
        #the edges of the combined nodes are now just the edges of node1 and node2
        combinedNode.edges = node1.edges + node2.edges
        #for all other nodes if they pointed to node1 or node2 now need to point to the
        # combined node
        for node in self.nodes:
            node.refresh(node1,node2,combinedNode)



    def __repr__(self):
        #for debugging if graph printed it should print a list of the nodes
        return str(self.nodes)







class Edge:
    #Edge class which houses the weight and the two nodes it connects
    def __init__(self,node1,node2,edge):
        self.weight = edge
        self.node1 = node1
        self.node2 = node2
        self.nodes = [node1,node2]

    def refresh(self,node1,node2,combinedNode):
        #if nodes being combined this replaces the pointer to node1 or node2 with the combined
        #node
        if self.node1 == node1 or self.node1 == node2:
            self.node1 == combinedNode
            self.nodes[0] = combinedNode
        if self.node2 == node1 or self.node2 == node2:
            self.node2 = combinedNode
            self.nodes[1] = combinedNode

    def otherNode(self,nodeGot):
        #if want to know from one node where it points to this returns
        #the other node
        if nodeGot == self.node1:
            return self.node2
        return self.node1

    def __repr__(self):
        #for debugging prints the first node and the second node and then the weight
        return f"{self.node1.getName()} to {self.node2.getName()} with weight {self.weight}"

class Node:
    def __init__(self,name,weight):
        self.name = name
        self.weight = weight
        self.edges = []

    def refresh(self,node1,node2,combinedNode):
        #looks at all edges and if node1 or node2 in edges
        #replaces them with the combinedNode
        for edge in self.edges:
            edge.refresh(node1,node2,combinedNode)


    def getEdge(self,node):
        #Find a certain edge connecting to a certain node
        for edge in self.edges:
            if node in edge.nodes:
                return edge

    def removeEdgesWasted(self,nodes):
        #removes an edge if now no longer connects to valid node
        edgesToRemove = []
        for edge in self.edges:
            if edge.node1 not in nodes or edge.node2 not in nodes:
                edgesToRemove.append(edge)
        for edge in edgesToRemove:
            self.edges.remove(edge)


    def addEdge(self,edge):
        #adds an edge to its array of edges
        self.edges.append(edge)

    def getName(self):
        #function to encapsulate self.name
        return self.name

    def connectedTo(self):
        #returns all children node connecting to
        children = []
        for edge in self.edges:
            children.append(edge.otherNode(self))
        return children

    def connectedToForwards(self):
        #returns all children node connecting to
        children = []
        for edge in self.edges:
            if edge.node1 == self:
                children.append(edge.otherNode(self))
        return children

    def numChildren(self):
        #each edge connects to a child and therefore length of
        #edges is the number of children
        return len(self.edges)

    def __repr__(self):
        #for debugging to print the node it first prints the name
        #then the weight and then a list of its edges
        return f"Name : {self.name} Weight : {self.weight} Edges : {self.edges}"

class DynamicAnalyses:
    #Class for dynamic analyses
    def __init__(self,script,Excel,funcs,saveLocation,Database):
        self.database = Database
        self.saveLocation = saveLocation
        #First inputs must be run on the code and data collected
        self.InputsRun = InputRunning(script,Excel,funcs)
        self.excel = Excel
        #data collected from the InputsRunC Class
        self.path = self.InputsRun.getTimingInfo()
        self.PathsUsed = self.InputsRun.getdDictTimes()
        self.FuncCallSheet = self.InputsRun.FuncCallSheet
        self.dictNumRanTimes = self.InputsRun.numTimesCalledDictarr[0]
        self.TimesTaken = self.InputsRun.timesTaken
        self.Funcs = funcs
        self.numFuncsCalled = self.InputsRun.numFuncsCalled
        self.times = self.calcTimeSpentInFuncs(self.FuncCallSheet,funcs)
        #Graph created
        self.DyanmicGraph()
        #Outputs from inpts saved in excel
        self.FillOutput(self.InputsRun.SavedOutputs,Excel)

    def FillOutput(self,outputs,Excel):

        #### Objective 16,17

        #The colour to highlight any outputs that are wrong
        redFill = ex.styles.PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        #Load in spreadsheet
        book = ex.load_workbook(Excel)
        try:
            sheet2 = book['Outputs']
        except:
            book.create_sheet('Outputs')
            sheet2 = book['Outputs']
        counter = 0
        #loop through all batch of outputs
        for index,outputlist in enumerate(outputs):
            counter += 2
            #loop through each actual output and add to spreadsheet
            for index2,output in enumerate(outputlist):
                #method to acces the spreadsheet as a 2d array
                cell = sheet2.cell(row = index2+2,column = counter)
                cell.value = output
                #if output does not match wanted output it is highlighted red
                prevCell =  sheet2.cell(row = index2+2,column = counter-1)
                if cell.value != prevCell.value:
                    cell.fill = redFill
        #spreadsheet saved
        book.save(Excel)

    def calcTimeSpentInFuncs(self,FuncCallSheet,funcs):

        #### Objective 25

        ##### Unravelling the matrix

        #Algorithm to calculate the time spent solely in one function without it being in another
        #Creating a dictionary of the final times which will then be added to over the course of the function
        times = {i.getName() : 0 for i in funcs}
        times["module"] = 0
        start = time.time()
        #Func call sheet looped through which is a 2d array containing one array for each runtime
        for index,q in enumerate(FuncCallSheet):
            numFuncsHere = self.numFuncsCalled[index]
            numFuncsConcated = 0
            l = np.array(q)
            counter = 1
            indexStarted = 0
            timeToSub = 0
            #FuncLookingAt always the last function to have been seen while looping through the array
            FuncLookingAt = l[0][0]
            timeStarted = float(l[0][1])
            while True:
                #Check if function and not just to show some amount of time
                #Always keeps track of last function come into contact with and looking for a time comes into closing
                #of same function without any other functions being called inbtween
                if l[counter][0] != "NOTFUNC" and l[counter][0] == FuncLookingAt:
                    numFuncsConcated += 1
                    timeFinished = float(l[counter][1])
                    times[l[counter][0].split()[0]] += ((timeFinished - timeStarted) -  timeToSub)
                    timeToSub = 0
                    #Array is concatenated with everything inbetween function call and finish being replaced
                    # with just one time to show how long that bit took
                    l = np.concatenate((l[0:indexStarted], [["NOTFUNC",timeFinished - timeStarted]], l[counter+1:]))
                    counter = 0
                    #Skips over times to find next function call
                    while counter < len(l) and l[counter][0] == "NOTFUNC":
                        counter += 1
                    if counter >= len(l):
                        #if end reached then list has been fully concatenated and while loop can be exited
                        break
                    #as concateneted counter returns to the start
                    indexStarted = counter
                    timeToSub = 0
                    FuncLookingAt = l[counter][0]
                    timeStarted = float(l[counter][1])
                #if could not match and is a function then that becomes the new FuncLookingAt
                elif l[counter][0] != "NOTFUNC":
                    FuncLookingAt = l[counter][0]
                    indexStarted = counter
                    timeStarted = float(l[counter][1])
                    timeToSub = 0
                #if it was just a time FuncLookingAt stays the same but total time increases by the time in the item
                else:
                    timeToSub += float(l[counter][1])
                counter += 1
                if counter >= len(l):
                    break
        return times




    def DyanmicGraph(self):

        #### Objective 27

        #Dynamic graph created
        DG = nx.DiGraph()
        for i in self.PathsUsed:
            DG.add_edges_from([(i[0] + '\n' + str(round(self.times[i[0]],1)),i[1] + '\n' + str(round(self.times[i[1]],1)))])
        for i in self.dictNumRanTimes:
            DG.add_weighted_edges_from([(i[0] + '\n' + str(round(self.times[i[0]],1)),i[1] + '\n' + str(round(self.times[i[1]],1)),round(self.dictNumRanTimes[i],1))])

        self.database.saveGraph(DG,"DynamicFuncInteracting",self.saveLocation)
        self.DynamicFuncInteracting = DG

        #From data the graph class is also created
        self.myTree = Tree()
        for i in self.dictNumRanTimes:
            self.myTree.createPath(i[0],round(self.times[i[0]],1),i[1],round(self.times[i[1]],1),round(self.dictNumRanTimes[i],1))
        #reduce run on graph class to make it easier to view
        self.myTree.reduce()
        book = ex.load_workbook(self.excel)
        #New sheet to save info created
        try:
            sheet2 = book["Dynamic Data"]
        except:
            book.create_sheet("Dynamic Data")
            sheet2 = book["Dynamic Data"]
        sheet2.cell(row = 1,column = 1).value = "isCyclic"
        isCyclic = self.myTree.isCyclic()
        sheet2.cell(row = 2,column = 1).value = str(isCyclic)
        if not isCyclic:
            module = self.myTree.getNode("module")
            cyclomComplex = self.myTree.cyclomComplex(module)
            sheet2.cell(row = 1,column = 2).value = "cyclomComplex"
            sheet2.cell(row = 2,column = 2).value = str(cyclomComplex)
        book.save(self.excel)
        #New reduced graph also saved
        self.myTree.save(self.saveLocation,self.database)

class InputRunning():
    #class that runs the input and dynamically collects data
    def __init__(self,script,excelFile,Funcs):

        self.excelFile = excelFile
        self.Funcs = Funcs
        self.script = script
        #If no input data code just run 10 times

        #### Objective 12

        self.inputData = [[None for j in range(10)] for i in range(10)]
        #If there is an excel file the input data collected
        if excelFile is not None:
            self.inputData = []
            book = ex.load_workbook(excelFile)
            sheet1 = book['Sheet1']
            col = 1
            while True:
                cell = sheet1.cell(row = 1,column = col)
                #if cell value none then end of all inputs reached
                if cell.value is None:
                    break
                #placeholder contains all inputs for one batch and then appended to the inputData
                placeholder = []
                ro = 1
                while True:
                    cell = sheet1.cell(row = ro,column = col)
                    #if cell value None then end of this input batch reached
                    if cell.value is None:
                        break
                    #to input blank user must write "" so that it can be distuigshed from the end which is nothing
                    if cell.value == '""':
                        placeholder.append("")
                    else:
                        placeholder.append(cell.value)
                    ro += 1
                self.inputData.append(placeholder)
                col += 1
        #when inputs collected code is run
        self.RunInputs(self.script,self.inputData)

    def getTimingInfo(self):
        #encapsulating self.paths
        return self.paths

    def getdDictTimes(self):
        #encapsulating dictTimesArr
        return self.dictTimesArr[0]

    def funcSorting(self,script,funcs):

        #### Objective 24,26

        script2 = ""
        for lineNum,line in enumerate(script.splitlines()):
            for i in funcs:
                if i.getName() in line:
                    numAdded = 0
                    #The recursive regex command to find all function calls for a certain function

                    #### Recursive Regex

                    for index,m in enumerate(regex.finditer(i.getName() + r'(\(((((?1)|[^\(\)])|(\".*\"))|(\'.*\'))*\))', line)):
                        end = m.span()[0] + len(i.getName()) + 2

                        if line.lstrip()[0:4] == "def ":
                            #Adds the arguments to all functions to accept to extra arguements, what function called it and This
                            #function
                            if i.getNumParameter() > 0:
                                line = line[0:end-1] + "FuncName,ThisFunc," + line[end-1::]
                            else:
                                line = line[0:end-1] + "FuncName,ThisFunc" + line[end-1::]
                        else:
                            #Works out which function a function call is in
                            #it does this by getting the range of lines each function is in
                            #and getting the one that has its start closes to the function call
                            minlines = 100000
                            winningFunc = 0
                            for z in funcs:
                                if lineNum > z.getLineNums()[0] and lineNum <= z.getLineNums()[1]:
                                    numLines = z.getLineNums()[1] - z.getLineNums()[0]
                                    if numLines < minlines:
                                        minlines = numLines
                                        winningFunc = z
                            #if no function around call then module is calling it
                            if winningFunc == 0:
                                winningFunc = f"'module','{i.getName()}'"
                            #winning function name and name of function is put in as two extra parameters to the start
                            #of the function call
                            else:
                                winningFunc = "'" + winningFunc.getName() + "'" +"," + "'" + i.getName() + "'"
                            if i.getNumParameter() > 0:
                                line = line[0:end-1 + numAdded]  + winningFunc + ',' + line[numAdded + end-1::]
                                numAdded += 1 + len(winningFunc)
                            else:
                                line = line[0:end-1 + numAdded] + winningFunc + line[numAdded + end-1::]
                                numAdded += len(winningFunc)


            script2 = script2 + line + "\n"
        return script2



    def RunInputs(self,script,inputs):

        #### Objective 13,14,15,24,26

        #function that runs the inputs
        self.prints = print
        #puts the two extra parameters to all the functions
        script = self.funcSorting(script,self.Funcs)
        #Regex to add the decorator to the top of all the funtion definitions
        AddingDecor = regex.compile(r"(( )*def )")
        scriptWithInputs = AddingDecor.sub("\\2\\2\\2\\2@funcDecorator\n\\1",script)
        dictOfName = {i.getName():0 for i in self.Funcs}
        dictOfName["module"] = 0
        #text that is added to the code which includes the decorator function definition
        #and the definition of some variables
        decorString = """import inspect
import time
times = []
numFuncs = 0
def funcDecorator(a_func):
    def wrapTheFunction(*args, **kwargs):
        global numFuncs
        numFuncs += 1
        FuncName = args[1] + " " + str(DictNames[args[1]])
        DictNames[args[1]] += 1
        #adding start timestamp to the array
        called.append((FuncName,time.time()))
        placeholder = a_func(*args, **kwargs)
        #adding finishing timestamp to the array
        called.append((FuncName,time.time()))

        #### Hash Table

        dicts.add((args[0],args[1]))
        #incremeting number of times this func called another by one
        if (args[0],args[1]) in numTimesCalledDict:
            numTimesCalledDict[(args[0],args[1])] += 1
        else:
            numTimesCalledDict[(args[0],args[1])] = 1
        return placeholder
    return wrapTheFunction"""

        decorString = "DictNames = " + str(dictOfName) + "\n" + decorString
        #adding decorstring to script that will run
        scriptWithInputs = decorString + "\n" + scriptWithInputs
        strtest = """SavedOutputs = []
dicts = set()
numTimesCalledDict = {}
#function to save variables
def saveToVariable(toSave):
    SavedOutputs.append(toSave)
#function to print is replaced with a function that will save whatever
#was meant to be printed
print = saveToVariable
def getNextInputXYZ(*args,**kwargs):
    return inputsxyzxyz.__next__()
input = getNextInputXYZ"""
        exec(strtest,globals())
        global numTimesCalledDict
        global dicts
        self.paths = []
        self.FuncCallSheet = []
        self.timesTaken = []
        self.numFuncsCalled = []
        self.numTimesCalledDictarr = []
        self.dictTimesArr = []
        self.SavedOutputs = []
        #for each batch of inputs code is run
        if len(inputs) == 0:
            inputs = [[1]]
        for input in inputs:
            #this specific input must be added as generator to the start of the code
            strToExec = "inputsxyzxyz = " + str(input)
            strToExec2 = "inputsxyzxyz = (item for item in inputsxyzxyz)"
            totString = strToExec + "\n" + strToExec2 + "\n" + scriptWithInputs
            starttime = time.time()
            totString = "import time" + "\n" + "called = []" + "\n" + "called.append(('module 1',time.time()))" + "\n" + totString + "\n" + "called.append(('module 1',time.time()))"
            #code is actually run
            exec(totString,globals())
            self.timesTaken.append(time.time()-starttime)
            global times
            #data is recorded and saved to the classes variables
            self.paths.append(times)
            self.FuncCallSheet.append(called)
            self.numFuncsCalled.append(numFuncs)
            self.numTimesCalledDictarr.append(dict(numTimesCalledDict))
            self.dictTimesArr.append(set(dicts))
            numTimesCalledDict = {}
            dicts = set()
            SavedOutputs.append("SPLIT")
        self.SavedOutputs = self.splitOutputs(SavedOutputs)

    def splitOutputs(self,SavedOutputs):
        #saved outputs is split up so that for each set of inputs
        #we have the corresponding set
        self.SavedOutputs = []
        placeholder = []
        for item in SavedOutputs:
            if item != "SPLIT":
                placeholder.append(item)
            else:
                self.SavedOutputs.append(placeholder)
                placeholder = []
        return self.SavedOutputs





class GUI:
    #GUI Class
    def __init__(self):
        self.ExcelFile = None
        self.saveLocation = None
        self.root = Tk()
        frame = Frame(self.root)
        frame.pack()
        self.root.title("Software Analyser")
        self.root.iconbitmap("GraphSquareSmall.ico")
        #Creating the buttons
        bottomframe = Frame(self.root)
        CodeFile = Button(frame, text = 'Find Code File', fg='brown',width = '25', command = self.askopenfile)
        CodeFile.pack( side = TOP )
        ExcelFile = Button(frame, text = 'Find Excel File', fg='brown',width = '25', command = self.askopenfileExcel)
        ExcelFile.pack( side = TOP )
        SaveLocation = Button(frame, text = 'Which folder to save?', fg='brown',width = '25', command = self.askSaveInfo)
        SaveLocation.pack( side = TOP )
        bottomframe.pack( side = BOTTOM )
        RunButton = Button(frame, text = 'Run Analyses', fg ='red',width = '25', command = self.run)
        RunButton.pack( side = TOP)


        self.root.mainloop()

    def run(self):
        #Creating main class when run button pressed
        messagebox.showinfo("Analyses started","The analyses of your code has begun")
        mains = Main(self.CodeFile,self.saveLocation,self.ExcelFile)
        messagebox.showinfo("Analyses complete","The analyses of your code is completed please view all files in the file you saved them")

    def askopenfile(self):

            #### Objective 1

            #Get code location
            self.CodeFile = filedialog.askopenfile(mode = 'r').name

    def askopenfileExcel(self):

            #### Objective 2

            #get Excel location
            self.ExcelFile = filedialog.askopenfile(mode = 'r').name

    def askSaveInfo(self):

            #### Objective 3

            #get save location information
            self.saveLocation = filedialog.askdirectory()


Gprint = print
sys.setrecursionlimit(3000)
myGUI = GUI()
